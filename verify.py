#!/usr/bin/env python3
"""Verification harness — returns non-zero on ANY failure.

Deterministic, offline checks (static fixtures) plus a pytest run.  Covers:
  1. Base reconciliation: index == 1000.00 on the base date.
  2. Divisor continuity: level immediately before == after every logged change,
     incl. a synthetic 3-stock "add a constituent" proof.
  3. Weights sum to 100% (post-cap) and no single weight exceeds the cap.
  4. No silent nulls: an un-sourced value surfaces as MISSING + anomaly.
  5. Idempotency: running the weekly job twice for one week yields one row.
  6. Unit + integration tests pass (pytest).
"""
from __future__ import annotations

import os
import subprocess
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from engine import config, indexmath
from engine.config import Constituent, Settings
from engine.datafetch import Quote, StaticDataProvider
from engine.fx import StaticFxProvider
from engine.state import EngineState
from engine.workbook import WorkbookManager
from engine.workflows import (
    apply_corporate_action, detect_anomalies, run_quarterly, run_weekly,
)
from engine.pipeline import fetch_all

TOL = 1e-6
HERE = os.path.dirname(os.path.abspath(__file__))


class Check:
    def __init__(self) -> None:
        self.failures: list[str] = []
        self.passes: list[str] = []

    def ok(self, name: str) -> None:
        self.passes.append(name)
        print(f"  [PASS] {name}")

    def fail(self, name: str, detail: str) -> None:
        self.failures.append(f"{name}: {detail}")
        print(f"  [FAIL] {name}: {detail}")


def _mk(ticker, price, ccy, shares, floatsh=None):
    return Quote(ticker=ticker, price=price, currency=ccy,
                 shares_outstanding=shares, float_shares=floatsh,
                 source="static", status="OK")


def _big_universe():
    """12-name synthetic universe: one name dominant (>10%) to exercise the cap,
    one name deliberately unresolvable (MISSING)."""
    consts = [Constituent(f"N{i}", f"T{i}", f"T{i}:X", "USD", "Fashion", "USA")
              for i in range(11)]
    consts.append(Constituent("Ghost", "GHOST", "G:X", "USD", "Fashion", "USA"))
    quotes = {"T0": _mk("T0", 100.0, "USD", 5_000_000, 4_500_000)}  # dominant
    for i in range(1, 11):
        quotes[f"T{i}"] = _mk(f"T{i}", 100.0, "USD", 500_000, 450_000)
    # GHOST intentionally absent -> MISSING
    data = StaticDataProvider(quotes)
    fx = StaticFxProvider({"USD": 1.0})
    return consts, data, fx


def _three_stock():
    consts = [
        Constituent("Alpha", "ALPHA", "A:X", "USD", "Fashion", "USA"),
        Constituent("Beta", "BETA.PA", "B:X", "EUR", "Luxury", "France"),
        Constituent("Gamma", "GAMMA.L", "G:X", "GBP", "Sportswear", "UK"),
    ]
    quotes = {
        "ALPHA": _mk("ALPHA", 100.0, "USD", 1_000_000, 900_000),
        "BETA.PA": _mk("BETA.PA", 50.0, "EUR", 2_000_000, 2_000_000),
        "GAMMA.L": _mk("GAMMA.L", 20.0, "GBP", 5_000_000, 4_000_000),
    }
    return consts, StaticDataProvider(quotes), StaticFxProvider(
        {"USD": 1.0, "EUR": 1.10, "GBP": 1.25})


def check_base_reconciliation(chk: Check) -> None:
    consts, data, fx = _three_stock()
    st = EngineState()
    wbm = WorkbookManager(os.path.join(HERE, ".verify_tmp.xlsx"))
    out = run_weekly(run_date="2026-01-05", data_provider=data, fx_provider=fx,
                     settings=Settings(), wbm=wbm, state=st,
                     constituents=consts, persist=False)
    if abs(out.index_level - 1000.0) <= TOL:
        chk.ok("1. Base reconciliation: index == 1000.00 on base date")
    else:
        chk.fail("1. Base reconciliation", f"level={out.index_level}")


def check_continuity(chk: Check) -> None:
    consts, data, fx = _three_stock()
    st = EngineState()
    wbm = WorkbookManager(os.path.join(HERE, ".verify_tmp.xlsx"))
    out = run_weekly(run_date="2026-01-05", data_provider=data, fx_provider=fx,
                     settings=Settings(), wbm=wbm, state=st,
                     constituents=consts, persist=False)
    level_before = out.index_level
    sum_before = indexmath.total_cap(st.last_caps_usd)
    new_cap = 3_000_000.0
    apply_corporate_action("Delta", "index-add", sum_before,
                           sum_before + new_cap, wbm=wbm, state=st,
                           persist=False)
    caps_after = dict(st.last_caps_usd, DELTA=new_cap)
    level_after = indexmath.index_level(caps_after, st.current_divisor)
    if abs(level_after - level_before) <= TOL:
        chk.ok("2a. Divisor continuity: add-a-constituent level unchanged "
               f"({level_before:.6f} == {level_after:.6f})")
    else:
        chk.fail("2a. Divisor continuity (synthetic add)",
                 f"{level_before} != {level_after}")

    # Quarterly proposal continuity.
    st2 = EngineState()
    wbm2 = WorkbookManager(os.path.join(HERE, ".verify_tmp2.xlsx"))
    run_weekly(run_date="2026-01-05", data_provider=data, fx_provider=fx,
               settings=Settings(), wbm=wbm2, state=st2, constituents=consts,
               persist=False)
    quotes2 = {
        "ALPHA": _mk("ALPHA", 100.0, "USD", 1_000_000, 900_000),
        "BETA.PA": _mk("BETA.PA", 50.0, "EUR", 2_400_000, 2_400_000),
        "GAMMA.L": _mk("GAMMA.L", 20.0, "GBP", 5_000_000, 4_000_000),
    }
    p = run_quarterly(run_date="2026-03-20", data_provider=StaticDataProvider(quotes2),
                      fx_provider=fx, settings=Settings(), wbm=wbm2, state=st2,
                      constituents=consts, persist=False)
    if abs(p.level_after - p.level_before) <= TOL:
        chk.ok("2b. Divisor continuity: quarterly rebalance level continuous")
    else:
        chk.fail("2b. Quarterly continuity",
                 f"{p.level_before} != {p.level_after}")


def check_weights(chk: Check) -> None:
    consts, data, fx = _big_universe()
    run = fetch_all(consts, "2026-01-05", data, fx, Settings())
    caps = run.caps_usd
    cap = 0.10
    weights = indexmath.compute_weights(caps, cap=cap)
    total = sum(weights.values())
    mx = max(weights.values())
    if abs(total - 1.0) <= TOL and mx <= cap + TOL:
        chk.ok(f"3. Weights sum to 100% post-cap (sum={total:.8f}) and "
               f"max weight {mx:.6f} <= cap {cap}")
    else:
        chk.fail("3. Weights/cap", f"sum={total}, max={mx}, cap={cap}")


def check_no_silent_nulls(chk: Check) -> None:
    consts, data, fx = _big_universe()
    st = EngineState()
    wbm = WorkbookManager(os.path.join(HERE, ".verify_tmp.xlsx"))
    out = run_weekly(run_date="2026-01-05", data_provider=data, fx_provider=fx,
                     settings=Settings(), wbm=wbm, state=st,
                     constituents=consts, persist=False)
    ghost_missing = out.n_missing >= 1 and any(
        a.kind == "MISSING" for a in out.anomalies)
    audit = wbm.read_rows("Audit")
    ghost_row = next((r for r in audit if r["ticker"] == "GHOST"), None)
    surfaced = ghost_row is not None and ghost_row["status"] == "MISSING" \
        and ghost_row["cap_usd"] in (None,)
    if ghost_missing and surfaced:
        chk.ok("4. No silent nulls: unresolved name is MISSING + anomaly "
               "(never 0/guessed)")
    else:
        chk.fail("4. No silent nulls",
                 f"n_missing={out.n_missing}, ghost_row={ghost_row}")


def check_idempotency(chk: Check) -> None:
    consts, data, fx = _three_stock()
    st = EngineState()
    wbm = WorkbookManager(os.path.join(HERE, ".verify_tmp.xlsx"))
    run_weekly(run_date="2026-01-05", data_provider=data, fx_provider=fx,
               settings=Settings(), wbm=wbm, state=st, constituents=consts,
               persist=False)
    out2 = run_weekly(run_date="2026-01-05", data_provider=data, fx_provider=fx,
                      settings=Settings(), wbm=wbm, state=st,
                      constituents=consts, persist=False)
    rows = wbm.read_rows("Weekly")
    if out2.skipped_duplicate and len(rows) == 1:
        chk.ok("5. Idempotency: weekly run twice for one week -> one row")
    else:
        chk.fail("5. Idempotency",
                 f"skipped={out2.skipped_duplicate}, rows={len(rows)}")


def check_live_workbook(chk: Check) -> None:
    """Validate the REAL artifact (if present) — not just synthetic fixtures."""
    from datetime import date as _date
    from openpyxl import load_workbook
    from engine.config import DEFAULT_SETTINGS

    path = config.WORKBOOK_PATH
    if not os.path.exists(path):
        print("  [SKIP] 7. Live workbook checks (no Fashion50_Index.xlsx yet — "
              "run backfill.py)")
        return
    wb = load_workbook(path, data_only=True)

    def rows(sheet):
        ws = wb[sheet]
        hdr = [c.value for c in ws[1]]
        out = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None or all(v is None for v in r):
                continue
            out.append({h: r[i] if i < len(r) else None
                        for i, h in enumerate(hdr)})
        return out

    weekly = rows("Weekly")
    # 7a. multi-week history (>= ~1 year of weekly points)
    if len(weekly) >= 50:
        chk.ok(f"7a. Live: Weekly time series has {len(weekly)} rows (>=50)")
    else:
        chk.fail("7a. Live weekly rows", f"only {len(weekly)} (run backfill.py)")

    # 7b. dates strictly increasing, weekly cadence
    dates = []
    ok_dates = True
    for r in weekly:
        try:
            dates.append(_date.fromisoformat(str(r["run_date"])))
        except Exception:
            ok_dates = False
    gaps_ok = all((dates[i] - dates[i - 1]).days <= 14 and dates[i] > dates[i - 1]
                  for i in range(1, len(dates)))
    if ok_dates and gaps_ok and dates == sorted(dates):
        chk.ok("7b. Live: dates strictly increasing at weekly cadence, no gaps")
    else:
        chk.fail("7b. Live date cadence", "non-monotonic or gap > 14 days")

    # 7c. every level > 0
    levels = [r["index_level"] for r in weekly if r.get("index_level") is not None]
    if levels and all(float(v) > 0 for v in levels):
        chk.ok(f"7c. Live: all {len(levels)} index levels > 0 (no silent zeros)")
    else:
        chk.fail("7c. Live positive levels", "found a non-positive level")

    # 7d. anchor ~ 1000 (index is rescaled to 1000 on the base/anchor date,
    # which may sit mid-series when history predates it)
    anchor = next((r for r in weekly
                   if str(r.get("run_date")) >= "2025-01-01"), None)
    if anchor and abs(float(anchor["index_level"]) - 1000.0) <= 1e-2:
        chk.ok(f"7d. Live: index == 1000.00 on anchor ({anchor['run_date']})")
    else:
        chk.fail("7d. Live anchor",
                 f"level at anchor = {anchor.get('index_level') if anchor else 'n/a'}")

    # 7e. every logged divisor change is continuous
    dl = rows("Divisor_Log")
    cont = True
    for r in dl:
        b, a = r.get("index_level_before"), r.get("index_level_after")
        if b is None or a is None or abs(float(b) - float(a)) > 1e-3:
            cont = False
    if cont:
        chk.ok(f"7e. Live: all {len(dl)} divisor changes continuous "
               "(level before == after)")
    else:
        chk.fail("7e. Live divisor continuity", "a logged change jumped the level")

    # 7f. current post-cap weights sum to 100% and <= cap
    audit = rows("Audit")
    caps = {r["ticker"]: float(r["cap_usd"]) for r in audit
            if r.get("cap_usd") and str(r.get("status")) == "OK"}
    if caps:
        cap = DEFAULT_SETTINGS.effective_cap
        w = indexmath.compute_weights(caps, cap=cap)
        tot, mx = sum(w.values()), max(w.values())
        if abs(tot - 1.0) <= 1e-6 and (cap is None or mx <= cap + 1e-6):
            chk.ok(f"7f. Live: current weights sum to 100% and max {mx:.4f} "
                   f"<= cap {cap}")
        else:
            chk.fail("7f. Live weights", f"sum={tot}, max={mx}, cap={cap}")
    else:
        chk.fail("7f. Live weights", "no OK caps in Audit")

    n_missing = weekly[-1].get("n_missing") if weekly else None
    print(f"  [INFO] latest week n_missing = {n_missing} "
          "(excluded/unsourced names are flagged, never zeroed)")


def check_preflight(chk: Check) -> None:
    """Early-warning scan surfaces upcoming events (synthetic, offline)."""
    from datetime import date as _date
    from engine.config import Constituent
    from engine.preflight import StaticPreflightSignals, TickerSignals, scan
    from engine.state import EngineState

    consts = [Constituent("Alpha", "ALPHA", "A:X", "USD", "Fashion", "USA")]
    st = EngineState()
    st.last_shares = {"ALPHA": 1_000_000}
    sig = StaticPreflightSignals({
        "ALPHA": TickerSignals(latest_shares=1_200_000,       # +20% issuance
                               next_earnings=_date(2026, 8, 10))})
    items = scan(consts, st, sig, _date(2026, 8, 4))
    cats = {i.category for i in items}
    if {"SHARES_DRIFT", "EARNINGS", "INDEX_REVIEW"} <= cats:
        chk.ok("8. Pre-flight: forward scan flags share drift + earnings + "
               "review countdown before they hit the calc")
    else:
        chk.fail("8. Pre-flight scan", f"categories={cats}")


def check_pytest(chk: Check) -> None:
    print("  ... running pytest ...")
    res = subprocess.run(
        [sys.executable, "-m", "pytest", "-q"],
        cwd=HERE, capture_output=True, text=True)
    tail = (res.stdout or "").strip().splitlines()[-1:] or [""]
    if res.returncode == 0:
        chk.ok(f"6. pytest ({tail[0].strip()})")
    else:
        print(res.stdout[-2000:])
        chk.fail("6. pytest", f"exit={res.returncode} {tail[0].strip()}")


def _cleanup() -> None:
    for f in (".verify_tmp.xlsx", ".verify_tmp2.xlsx"):
        p = os.path.join(HERE, f)
        try:
            if os.path.exists(p):
                os.remove(p)
        except OSError:
            pass


def main() -> int:
    print("=" * 70)
    print("VGR Fashion 50 — verification")
    print("=" * 70)
    chk = Check()
    try:
        check_base_reconciliation(chk)
        check_continuity(chk)
        check_weights(chk)
        check_no_silent_nulls(chk)
        check_idempotency(chk)
        check_preflight(chk)
        check_live_workbook(chk)
        check_pytest(chk)
    finally:
        _cleanup()

    print("-" * 70)
    if chk.failures:
        print(f"RESULT: FAILED ({len(chk.failures)} failure(s), "
              f"{len(chk.passes)} passed)")
        for f in chk.failures:
            print(f"   - {f}")
        return 1
    print(f"RESULT: ALL GREEN ({len(chk.passes)} checks passed)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
