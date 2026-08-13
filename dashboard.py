#!/usr/bin/env python3
"""Render a self-contained HTML dashboard + JSON feed from the workbook.

Design: VGR house style (black, Barlow / Barlow Condensed, warm-taupe accent,
"// SECTION" kickers, sharp corners) with Economist chart discipline — each card
carries a kicker, a declarative headline, a subtitle (metric · unit · period),
and a source line; minimal gridlines; a single accent for the hero series.

Usage:
    python dashboard.py [--open]
"""
from __future__ import annotations

import datetime
import html
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook  # noqa: E402

from engine import config, indexmath  # noqa: E402
from engine.config import DEFAULT_SETTINGS, load_constituents, primary_segment  # noqa: E402

DASHBOARD_PATH = os.path.join(config.PROJECT_DIR, "Fashion50_Dashboard.html")
DATA_JSON_PATH = os.path.join(config.PROJECT_DIR, "fashion50_data.json")
BASE_DATE = "2025-01-01"
BASE_LABEL = "2 Jan 2025"


def _read_workbook():
    wb = load_workbook(config.WORKBOOK_PATH, data_only=True)

    def rows(sheet):
        ws = wb[sheet]
        headers = [c.value for c in ws[1]]
        out = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None or all(v is None for v in r):
                continue
            out.append({h: r[i] if i < len(r) else None
                        for i, h in enumerate(headers)})
        return out

    def wide(sheet):
        order, data = [], {}
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            hdr = [c.value for c in ws[1]]
            order = hdr[1:]
            for c in order:
                data[c] = []
            for r in ws.iter_rows(min_row=2, values_only=True):
                d = r[0]
                if d is None:
                    continue
                for j, c in enumerate(order, start=1):
                    v = r[j] if j < len(r) else None
                    if v is not None:
                        data[c].append((str(d), float(v)))
        return order, data

    _, prices = wide("Prices")
    seg_order, subidx = wide("SubIndices")
    return (rows("Weekly"), rows("Audit"), rows("Constituents"),
            rows("Unscheduled"), prices, seg_order, subidx)


def _fmt(x, nd=2):
    try:
        return f"{float(x):,.{nd}f}"
    except (TypeError, ValueError):
        return "—"


def _pct(x, nd=2):
    if x is None:
        return "—"
    return f"{x:+.{nd}f}%"


def _sign(v):
    if v is None:
        return "flat"
    return "up" if v >= 0 else "down"


def _changes(series) -> dict:
    s = [(d, v) for d, v in series if v is not None]
    out = {"latest": None, "yesterday": None, "l7d": None, "l30d": None,
           "mtd": None, "ytd": None}
    if not s:
        return out
    ld, lv = s[-1]
    out["latest"] = lv

    def base_days(n):
        target = datetime.date.fromisoformat(ld) - datetime.timedelta(days=n)
        b = None
        for d, v in s:
            if datetime.date.fromisoformat(d) <= target:
                b = v
            else:
                break
        return b

    def base_before(since_iso):
        b = None
        for d, v in s:
            if d < since_iso:
                b = v
            else:
                break
        return b

    if len(s) >= 2 and s[-2][1]:
        out["yesterday"] = (lv / s[-2][1] - 1) * 100
    for key, n in (("l7d", 7), ("l30d", 30)):
        b = base_days(n)
        if b:
            out[key] = (lv / b - 1) * 100
    y, m = ld[:4], ld[5:7]
    bm = base_before(f"{y}-{m}-01")
    if bm:
        out["mtd"] = (lv / bm - 1) * 100
    by = base_before(f"{y}-01-01")
    if by:
        out["ytd"] = (lv / by - 1) * 100
    return out


def _metric_tile(label, pct):
    return (f'<div class="kpi"><div class="k-label">{label}</div>'
            f'<div class="k-val {_sign(pct)}">{_pct(pct)}</div></div>')


def _cardhead(kicker, headline, dek):
    return (f'<div class="kicker">// {html.escape(kicker)}</div>'
            f'<div class="headline">{html.escape(headline)}</div>'
            f'<div class="dek">{html.escape(dek)}</div>')


def _segment_table(seg_order, subidx):
    body = []
    for seg in seg_order:
        m = _changes(subidx.get(seg, []))
        body.append(
            f"<tr><td><b>{html.escape(str(seg))}</b></td>"
            f"<td class='num'>{_fmt(m['latest'], 1)}</td>"
            f"<td class='num {_sign(m['yesterday'])}'>{_pct(m['yesterday'])}</td>"
            f"<td class='num {_sign(m['l7d'])}'>{_pct(m['l7d'])}</td>"
            f"<td class='num {_sign(m['l30d'])}'>{_pct(m['l30d'])}</td>"
            f"<td class='num {_sign(m['mtd'])}'>{_pct(m['mtd'])}</td>"
            f"<td class='num {_sign(m['ytd'])}'>{_pct(m['ytd'])}</td></tr>")
    return (
        "<table class='wt'><thead><tr><th>Segment</th>"
        "<th class='num'>Level</th><th class='num'>Yday</th><th class='num'>L7D</th>"
        "<th class='num'>L30D</th><th class='num'>MTD</th><th class='num'>YTD</th>"
        "</tr></thead><tbody>" + "".join(body) + "</tbody></table>")


def _constituents_table(recs):
    body = []
    for r in recs:
        nm = html.escape(str(r["name"]))
        if r.get("ir"):
            nm = (f'<a href="{html.escape(r["ir"])}" target="_blank" '
                  f'rel="noopener noreferrer">{nm}</a>')
        body.append(
            f"<tr><td class='r'>{r['rank']}</td>"
            f"<td>{nm}<div class='sub'>{html.escape(str(r['ticker']))}</div></td>"
            f"<td>{html.escape(str(r['segment']))}</td>"
            f"<td class='num'>{_fmt(r['price'])}<div class='sub'>{r['ccy']}</div></td>"
            f"<td class='num {_sign(r['yday'])}'>{_pct(r['yday'], 2)}</td>"
            f"<td class='num {_sign(r['l7d'])}'>{_pct(r['l7d'], 2)}</td>"
            f"<td class='num {_sign(r['ytd'])}'>{_pct(r['ytd'], 1)}</td>"
            f"<td class='num'>{_fmt(r['fullcap'], 1)}</td>"
            f"<td class='num'>{r['float']:.0f}%</td>"
            f"<td class='num'>{r['wt']:.2f}%</td></tr>")
    return (
        "<table class='wt'><thead><tr>"
        "<th class='r'>#</th><th>Company</th><th>Segment</th>"
        "<th class='num'>Price</th><th class='num'>Yday</th><th class='num'>L7D</th>"
        "<th class='num'>YTD</th><th class='num'>Cap $bn</th>"
        "<th class='num'>Float</th><th class='num'>Weight</th>"
        "</tr></thead><tbody>" + "".join(body) + "</tbody></table>")


def _prepare(weekly, audit, consts, prices):
    settings = DEFAULT_SETTINGS
    series = [(str(r["run_date"]), round(float(r["index_level"]), 4))
              for r in weekly if r.get("index_level") is not None]
    seg_of = {c["yahoo_ticker"]: c.get("segment", "?") for c in consts}
    name_of = {c["yahoo_ticker"]: c.get("name") for c in consts}
    ir_of = {c.yahoo_ticker: c.ir_url for c in load_constituents()}
    ccy_of = {r["ticker"]: r.get("currency") for r in audit}
    ff_of = {r["ticker"]: (r.get("float_factor") or 1.0) for r in audit}
    caps = {r["ticker"]: float(r["cap_usd"]) for r in audit
            if r.get("cap_usd") and r.get("status") == "OK"}
    total_cap = sum(caps.values()) or 1.0
    weights = indexmath.compute_weights(caps, cap=settings.effective_cap)
    ranked = sorted(weights, key=weights.get, reverse=True)

    recs = []
    for i, t in enumerate(ranked, 1):
        ch = _changes(prices.get(t, []))
        ff = ff_of.get(t, 1.0) or 1.0
        recs.append({
            "rank": i, "ticker": t, "name": name_of.get(t, t),
            "ir": ir_of.get(t, ""), "segment": primary_segment(seg_of.get(t)),
            "price": ch["latest"], "ccy": ccy_of.get(t, ""),
            "yday": ch["yesterday"], "l7d": ch["l7d"], "ytd": ch["ytd"],
            "fullcap": caps[t] / ff / 1e9, "float": ff * 100,
            "raw": caps[t] / total_cap * 100, "wt": weights[t] * 100,
        })
    idx_metrics = _changes([(d, lv) for d, lv in series])
    return series, recs, idx_metrics, name_of, ranked


def build_html() -> str:
    weekly, audit, consts, unsched, prices, seg_order, subidx = _read_workbook()
    series, recs, m, name_of, ranked = _prepare(weekly, audit, consts, prices)
    latest = m["latest"] or 0.0
    ytd = m["ytd"]
    as_of = series[-1][0] if series else "—"
    source = (f"Source: VGR · Yahoo Finance (prices), ECB via Frankfurter (FX) "
              f"· daily · as of {as_of}")

    def finding(v):
        if v is None:
            return "flat year to date"
        return (f"up {v:.1f}% year to date" if v >= 0
                else f"down {abs(v):.1f}% year to date")
    lead_seg = seg_order[0] if seg_order else None
    seg_ytd = {s: _changes(subidx.get(s, []))["ytd"] for s in seg_order}
    best = max((s for s in seg_order if seg_ytd.get(s) is not None),
               key=lambda s: seg_ytd[s], default=None)
    worst = min((s for s in seg_order if seg_ytd.get(s) is not None),
                key=lambda s: seg_ytd[s], default=None)

    sel_prices = {t: [[d, p] for d, p in prices.get(t, [])] for t in ranked
                  if prices.get(t)}
    sel_names = {t: name_of.get(t, t) for t in sel_prices}
    idx_series = [[d, lv] for d, lv in series]
    start_year = int(series[0][0][:4]) if series else 2020
    sel_order = sorted((t for t in ranked if t in sel_prices),
                       key=lambda t: (name_of.get(t, t) or "").lower())
    default_t = next((t for t in ranked if t in sel_prices), None)
    options = "".join(
        f'<option value="{t}"{" selected" if t == default_t else ""}>'
        f'{html.escape(str(name_of.get(t, t)))}</option>' for t in sel_order)

    horizons = [("Since 2025", "2025-01-01")]
    if start_year <= 2023:
        horizons.append(("Since 2023", "2023-01-01"))
    horizons.append((f"Since {start_year}", series[0][0] if series else "2020-01-01"))
    hbtns = "".join(
        f'<button class="hbtn{" active" if i == 0 else ""}" '
        f'data-start="{s}">{lab}</button>'
        for i, (lab, s) in enumerate(horizons))

    kpis = (
        f'<div class="kpi hero"><div class="k-label">Index level</div>'
        f'<div class="k-val big">{_fmt(latest, 1)}</div>'
        f'<div class="k-sub">1,000 = {BASE_LABEL}</div></div>'
        + _metric_tile("Yesterday", m["yesterday"])
        + _metric_tile("L7D", m["l7d"])
        + _metric_tile("L30D", m["l30d"])
        + _metric_tile("MTD", m["mtd"])
        + _metric_tile("YTD", m["ytd"]))

    seg_dek = "Float-adjusted market cap · USD · 1,000 = " + BASE_LABEL
    if best and worst and best != worst:
        seg_dek = (f"{best} leads ({seg_ytd[best]:+.0f}% YTD), "
                   f"{worst} lags ({seg_ytd[worst]:+.0f}% YTD)")

    head = f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>VGR 50 — Largest Lifestyle Companies</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Barlow:wght@400;500;600;700&family=Barlow+Condensed:wght@600;700;800&display=swap" rel="stylesheet">
<style>
:root {{ --bg:#000; --card:#0c0c0d; --line:rgba(255,255,255,.13);
  --ink:#E5E7EB; --muted:rgba(255,255,255,.46); --accent:#C5C1B9;
  --up:#57cf9a; --down:#e8695f; --grid:rgba(255,255,255,.08); }}
* {{ box-sizing:border-box; }}
body {{ margin:0; background:var(--bg); color:var(--ink);
  font-family:"Barlow",-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
  font-size:15px; line-height:1.5; -webkit-font-smoothing:antialiased; }}
.wrap {{ max-width:1140px; margin:0 auto; padding:34px 22px 72px; }}
.cond {{ font-family:"Barlow Condensed","Barlow",sans-serif; }}
.kicker {{ font-family:"Barlow Condensed",sans-serif; text-transform:uppercase;
  letter-spacing:.16em; color:var(--accent); font-size:12px; font-weight:700; }}
header .kicker {{ font-size:13px; }}
h1 {{ font-family:"Barlow Condensed",sans-serif; text-transform:uppercase;
  font-weight:800; font-size:44px; line-height:1.02; letter-spacing:.01em;
  margin:6px 0 8px; }}
header .dek {{ color:var(--muted); font-size:14px; max-width:70ch; }}
header .finding {{ font-family:"Barlow Condensed",sans-serif; text-transform:uppercase;
  font-weight:700; font-size:17px; letter-spacing:.02em; margin-top:14px;
  color:var(--ink); }}
.headline {{ font-family:"Barlow Condensed",sans-serif; text-transform:uppercase;
  font-weight:700; font-size:19px; letter-spacing:.02em; margin:2px 0 3px; }}
.dek {{ color:var(--muted); font-size:13px; }}
.rule {{ height:2px; width:64px; background:var(--accent); margin:0 0 6px; }}
.kpis {{ display:grid; grid-template-columns:repeat(6,1fr); gap:1px;
  background:var(--line); border:1px solid var(--line); margin:24px 0; }}
.kpi {{ background:var(--bg); padding:16px 18px; }}
.k-label {{ font-family:"Barlow Condensed",sans-serif; color:var(--muted);
  font-size:12px; text-transform:uppercase; letter-spacing:.12em; }}
.k-val {{ font-size:26px; font-weight:600; margin-top:6px; font-variant-numeric:tabular-nums; }}
.k-val.big {{ color:var(--accent); }}
.k-sub {{ font-size:11px; color:var(--muted); margin-top:2px; }}
.up {{ color:var(--up); }} .down {{ color:var(--down); }}
.card {{ background:var(--card); border:1px solid var(--line); padding:22px 24px; margin:16px 0; }}
.chart {{ width:100%; height:auto; margin-top:8px; }}
.grid {{ stroke:var(--grid); stroke-width:1; }}
.baseline {{ stroke:var(--muted); stroke-width:1; stroke-dasharray:4 4; }}
.vline {{ stroke:var(--accent); stroke-width:1.1; stroke-dasharray:3 3; }}
.axis {{ fill:var(--muted); font-size:11px; font-family:"Barlow",sans-serif; }}
.hrow {{ display:flex; gap:0; margin:12px 0 4px; border:1px solid var(--line); width:fit-content; }}
.hbtn {{ background:transparent; color:var(--muted); border:0; border-right:1px solid var(--line);
  padding:6px 16px; font-family:"Barlow Condensed",sans-serif; text-transform:uppercase;
  letter-spacing:.08em; font-size:12px; font-weight:600; cursor:pointer; }}
.hbtn:last-child {{ border-right:0; }}
.hbtn.active {{ background:var(--accent); color:#000; }}
.source {{ color:var(--muted); font-size:11px; margin-top:14px; padding-top:9px;
  border-top:1px solid var(--line); }}
table.wt {{ width:100%; border-collapse:collapse; font-size:13px; margin-top:8px; }}
table.wt th {{ text-align:left; color:var(--muted); font-weight:600; font-size:11px;
  font-family:"Barlow Condensed",sans-serif; text-transform:uppercase; letter-spacing:.08em;
  padding:7px 8px; border-bottom:1px solid var(--line); }}
table.wt td {{ padding:7px 8px; border-bottom:1px solid rgba(255,255,255,.06); vertical-align:top; }}
table.wt tr:hover td {{ background:rgba(255,255,255,.02); }}
table.wt td.r {{ color:var(--muted); width:30px; text-align:right; font-variant-numeric:tabular-nums; }}
table.wt th.num, table.wt td.num {{ text-align:right; font-variant-numeric:tabular-nums; white-space:nowrap; }}
table.wt a {{ color:var(--ink); text-decoration:none; border-bottom:1px solid var(--accent); }}
table.wt a:hover {{ color:var(--accent); }}
.sub {{ color:var(--muted); font-size:11px; }}
select {{ background:#000; color:var(--ink); border:1px solid var(--line); border-radius:0;
  padding:8px 12px; font-family:"Barlow",sans-serif; font-size:14px; max-width:340px; margin-top:10px; }}
.cocap {{ color:var(--muted); font-size:13px; margin-top:8px; }}
footer {{ margin-top:34px; padding-top:16px; border-top:1px solid var(--line); }}
footer .dek {{ font-size:11px; }}
@media (max-width:860px) {{ .kpis {{ grid-template-columns:repeat(3,1fr); }} h1 {{ font-size:34px; }} }}
@media (max-width:520px) {{ .kpis {{ grid-template-columns:repeat(2,1fr); }} }}
</style></head>
<body><div class="wrap">
<header>
  <div class="rule"></div>
  <div class="kicker">// VGR INTELLIGENCE</div>
  <h1>VGR 50 — Largest Lifestyle Companies</h1>
  <div class="dek">A market-cap-weighted, float-adjusted index of the 50 largest listed fashion, luxury and sportswear companies. Priced daily in US dollars; 1,000 = {BASE_LABEL}.</div>
  <div class="finding">VGR 50 is {finding(ytd)}</div>
</header>
<div class="kpis">{kpis}</div>
<div class="card">
  {_cardhead("INDEX", "VGR 50 " + finding(ytd), "Daily · index level · USD · 1,000 = " + BASE_LABEL)}
  <div class="hrow">{hbtns}</div>
  <div id="idxchart"></div>
  <div class="source">{html.escape(source)}</div>
</div>
<div class="card">
  {_cardhead("SEGMENTS", "Segment sub-indices", seg_dek)}
{_segment_table(seg_order, subidx)}
  <div class="source">Each segment is its own float-adjusted sub-index, rebased to 1,000 on {BASE_LABEL}. {html.escape(source)}</div>
</div>
<div class="card">
  {_cardhead("CONSTITUENTS", "The 50 largest listed lifestyle companies", "Ranked by float-adjusted market cap · local share price · company names link to investor relations")}
{_constituents_table(recs)}
  <div class="source">Weight is the float-adjusted market-cap share; cap is full size (shares × price); float is the tradable share. {html.escape(source)}</div>
</div>
<div class="card">
  {_cardhead("COMPANY", "Single-company evolution", "Any constituent, rebased to 1,000 at " + BASE_LABEL + ", against the index")}
  <select id="co">{options}</select>
  <div id="cochart"></div>
  <div class="cocap" id="cocap"></div>
</div>
<footer>
  <div class="kicker">// VGR — VERY GOOD RETAIL</div>
  <div class="dek">Shares and float held at the current snapshot (a documented approximation); pre-{BASE_DATE[:4]} history is a reconstruction. Not investment advice.</div>
</footer>
</div>"""

    script = """
<script>
const IDXFULL = __IDXFULL__;
const BASE = __BASE__;
const PRICES = __PRICES__;
const NAMES = __NAMES__;
const W=1040,H=320,pad=48;

function drawIndex(startISO){
  const data = IDXFULL.filter(r=>r[0]>=startISO);
  if(data.length<2) return;
  const vals=data.map(r=>r[1]);
  let lo=Math.min.apply(null,vals), hi=Math.max.apply(null,vals);
  const span=(hi-lo)||1;
  const Y=v=>pad+(H-2*pad)*(1-(v-lo)/span);
  const poly=a=>a.map((r,i)=>((pad+(W-2*pad)*i/(a.length-1)).toFixed(1))+","+Y(r[1]).toFixed(1)).join(" ");
  const pts=poly(data);
  const area=pad.toFixed(1)+","+(H-pad).toFixed(1)+" "+pts+" "+(W-pad).toFixed(1)+","+(H-pad).toFixed(1);
  let grid="";
  for(const f of [0,.25,.5,.75,1]){const val=lo+span*f,yy=Y(val);
    grid+='<line x1="'+pad+'" y1="'+yy.toFixed(1)+'" x2="'+(W-pad)+'" y2="'+yy.toFixed(1)+'" class="grid"/>'+
    '<text x="'+(pad-8)+'" y="'+(yy+4).toFixed(1)+'" text-anchor="end" class="axis">'+val.toFixed(0)+'</text>';}
  let base1000="";
  if(lo<=1000&&1000<=hi){const yb=Y(1000);
    base1000='<line x1="'+pad+'" y1="'+yb.toFixed(1)+'" x2="'+(W-pad)+'" y2="'+yb.toFixed(1)+'" class="baseline"/>';}
  let vline="";
  let bi=data.findIndex(r=>r[0]>=BASE);
  if(bi>0){const xb=pad+(W-2*pad)*bi/(data.length-1);
    vline='<line x1="'+xb.toFixed(1)+'" y1="'+pad+'" x2="'+xb.toFixed(1)+'" y2="'+(H-pad)+'" class="vline"/>'+
    '<text x="'+(xb+4).toFixed(1)+'" y="'+(pad+12)+'" class="axis" fill="var(--accent)">'+BASE+' = 1000</text>';}
  let xlab="";
  for(const f of [0,.25,.5,.75,1]){const i=Math.round((data.length-1)*f);
    const x=pad+(W-2*pad)*i/(data.length-1);
    xlab+='<text x="'+x.toFixed(1)+'" y="'+(H-pad+22)+'" text-anchor="middle" class="axis">'+data[i][0]+'</text>';}
  const lx=pad+(W-2*pad), ly=Y(data[data.length-1][1]);
  document.getElementById("idxchart").innerHTML=
    '<svg viewBox="0 0 '+W+' '+H+'" class="chart">'+grid+base1000+vline+
    '<polygon points="'+area+'" fill="var(--accent)" fill-opacity="0.10"/>'+
    '<polyline points="'+pts+'" fill="none" stroke="var(--accent)" stroke-width="2"/>'+
    '<circle cx="'+lx.toFixed(1)+'" cy="'+ly.toFixed(1)+'" r="3.5" fill="var(--accent)"/>'+
    xlab+'</svg>';
}
document.querySelectorAll(".hbtn").forEach(b=>b.addEventListener("click",e=>{
  document.querySelectorAll(".hbtn").forEach(x=>x.classList.remove("active"));
  e.target.classList.add("active");
  drawIndex(e.target.dataset.start);
}));
drawIndex("__BASE__");

function drawCo(t){
  const raw=(PRICES[t]||[]).filter(r=>r[1]!=null);
  if(!raw.length) return;
  const base=raw[0][1];
  const comp=raw.map(r=>[r[0], r[1]/base*1000]);
  const idx=IDXFULL.filter(r=>r[0]>=BASE);
  const vals=comp.map(r=>r[1]).concat(idx.map(r=>r[1]));
  let lo=Math.min.apply(null,vals), hi=Math.max.apply(null,vals);
  const span=(hi-lo)||1;
  const Y=v=>pad+(H-2*pad)*(1-(v-lo)/span);
  const poly=a=>a.map((r,i)=>((pad+(W-2*pad)*i/(a.length-1)).toFixed(1))+","+Y(r[1]).toFixed(1)).join(" ");
  let grid="";
  for(const f of [0,.25,.5,.75,1]){const val=lo+span*f,yy=Y(val);
    grid+='<line x1="'+pad+'" y1="'+yy.toFixed(1)+'" x2="'+(W-pad)+'" y2="'+yy.toFixed(1)+'" class="grid"/>'+
    '<text x="'+(pad-8)+'" y="'+(yy+4).toFixed(1)+'" text-anchor="end" class="axis">'+val.toFixed(0)+'</text>';}
  let b1="";
  if(lo<=1000&&1000<=hi){const yb=Y(1000);
    b1='<line x1="'+pad+'" y1="'+yb.toFixed(1)+'" x2="'+(W-pad)+'" y2="'+yb.toFixed(1)+'" class="baseline"/>';}
  let xlab="";
  for(const f of [0,.25,.5,.75,1]){const i=Math.round((idx.length-1)*f);
    const x=pad+(W-2*pad)*i/(idx.length-1);
    xlab+='<text x="'+x.toFixed(1)+'" y="'+(H-pad+22)+'" text-anchor="middle" class="axis">'+idx[i][0]+'</text>';}
  document.getElementById("cochart").innerHTML=
    '<svg viewBox="0 0 '+W+' '+H+'" class="chart">'+grid+b1+
    '<polyline points="'+poly(idx)+'" fill="none" stroke="var(--muted)" stroke-width="1.5" stroke-dasharray="5 4"/>'+
    '<polyline points="'+poly(comp)+'" fill="none" stroke="var(--accent)" stroke-width="2.3"/>'+xlab+'</svg>';
  const last=comp[comp.length-1][1], pct=(last/1000-1)*100;
  document.getElementById("cocap").innerHTML=
    "<b style='color:var(--accent)'>"+NAMES[t]+"</b> rebased 1,000 -> "+last.toFixed(1)+
    " ("+(pct>=0?"+":"")+pct.toFixed(1)+"% since base). Dashed = the VGR 50 index.";
}
const sel=document.getElementById("co");
sel.addEventListener("change",e=>drawCo(e.target.value));
drawCo(sel.value);
</script>
</body></html>"""
    script = (script
              .replace("__IDXFULL__", json.dumps(idx_series))
              .replace("__PRICES__", json.dumps(sel_prices))
              .replace("__NAMES__", json.dumps(sel_names))
              .replace("__BASE__", json.dumps(BASE_DATE)))
    return head + script


def _r(x, nd=2):
    return round(x, nd) if isinstance(x, (int, float)) else None


def build_data() -> dict:
    weekly, audit, consts, unsched, prices, seg_order, subidx = _read_workbook()
    series, recs, m, name_of, ranked = _prepare(weekly, audit, consts, prices)
    segments = []
    for seg in seg_order:
        sm = _changes(subidx.get(seg, []))
        segments.append({
            "segment": seg, "level": _r(sm["latest"]),
            "chg_yesterday_pct": _r(sm["yesterday"]), "chg_l7d_pct": _r(sm["l7d"]),
            "chg_l30d_pct": _r(sm["l30d"]), "mtd_pct": _r(sm["mtd"]),
            "ytd_pct": _r(sm["ytd"]),
        })
    return {
        "generated_at": datetime.datetime.now(datetime.timezone.utc)
        .isoformat(timespec="seconds"),
        "base_date": BASE_DATE, "base_level": 1000.0,
        "history_start": series[0][0] if series else None,
        "latest_date": series[-1][0] if series else None,
        "latest_level": m["latest"],
        "index_change": {"yesterday": _r(m["yesterday"]), "l7d": _r(m["l7d"]),
                         "l30d": _r(m["l30d"]), "mtd": _r(m["mtd"]),
                         "ytd": _r(m["ytd"])},
        "n_constituents": len(recs),
        "index": [[d, lv] for d, lv in series],
        "segments": segments,
        "subindices": {seg: [[d, round(v, 3)] for d, v in subidx.get(seg, [])]
                       for seg in seg_order},
        "constituents": [{
            "rank": r["rank"], "ticker": r["ticker"], "name": r["name"],
            "ir_url": r["ir"], "segment": r["segment"],
            "price": _r(r["price"], 4), "currency": r["ccy"],
            "chg_yesterday_pct": _r(r["yday"]), "chg_l7d_pct": _r(r["l7d"]),
            "ytd_pct": _r(r["ytd"], 1), "full_cap_usd_bn": _r(r["fullcap"], 3),
            "float_pct": _r(r["float"], 1), "weight_pct": _r(r["wt"], 4),
        } for r in recs],
        "prices": {t: [[d, round(p, 4)] for d, p in prices.get(t, [])]
                   for t in ranked if prices.get(t)},
    }


def main(argv: list[str]) -> int:
    if not os.path.exists(config.WORKBOOK_PATH):
        print("No workbook found — run backfill.py first.")
        return 1
    with open(DASHBOARD_PATH, "w", encoding="utf-8") as fh:
        fh.write(build_html())
    print(f"Wrote {DASHBOARD_PATH}")
    with open(DATA_JSON_PATH, "w", encoding="utf-8") as fh:
        json.dump(build_data(), fh, ensure_ascii=False)
    print(f"Wrote {DATA_JSON_PATH}")
    if "--open" in argv:
        try:
            import subprocess
            subprocess.Popen(["cmd", "/c", "start", "chrome", DASHBOARD_PATH])
        except Exception:
            print("(could not auto-open Chrome; open the file manually)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
