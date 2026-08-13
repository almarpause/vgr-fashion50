"""Excel workbook I/O (openpyxl).

Owns the single artifact ``Fashion50_Index.xlsx`` with the four workflow
worksheets — Weekly, Quarterly, Annual, Unscheduled — plus supporting sheets
Constituents, FX, Alerts, Divisor_Log.  Row appends are schema-checked; the
Weekly sheet is de-duplicated by ISO week for idempotency.
"""
from __future__ import annotations

import os
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from . import config

# Sheet name -> ordered column headers.
SCHEMAS: dict[str, list[str]] = {
    "Weekly": [
        "run_date", "iso_week", "index_level", "weekly_return_%",
        "total_mcap_usd", "divisor_used", "n_ok", "n_missing", "anomaly_flag",
    ],
    "Quarterly": [
        "run_date", "ticker", "name", "old_weight_%", "new_weight_%",
        "weight_change_%", "old_shares", "new_shares",
        "old_divisor", "new_divisor", "biggest_mover", "status",
    ],
    "Annual": [
        "run_date", "ticker", "name", "action", "rank",
        "float_adj_cap_usd", "segment", "scope_ok", "status",
    ],
    "Unscheduled": [
        "effective_date", "company", "event_type", "detected_by",
        "sum_before_usd", "sum_after_usd", "old_divisor", "new_divisor",
        "status",
    ],
    "Constituents": [
        "name", "yahoo_ticker", "google_ticker", "trading_currency",
        "segment", "country_hq", "float_manual_review",
    ],
    "FX": [
        "run_date", "currency", "rate_to_usd", "source", "status",
    ],
    "Alerts": [
        "timestamp", "severity", "subject", "detail",
    ],
    "Divisor_Log": [
        "timestamp", "reason", "old_divisor", "new_divisor",
        "sum_before_usd", "sum_after_usd",
        "index_level_before", "index_level_after", "source_sheet",
    ],
    "Audit": [
        "run_date", "ticker", "name", "price_major", "currency",
        "shares_outstanding", "float_shares", "float_factor",
        "fx_rate_to_usd", "cap_usd", "price_source", "fx_source", "status",
    ],
    "Watchlist": [
        "scan_date", "ticker", "company", "category", "signal", "severity",
        "event_date", "days_out", "suggested_action", "source",
    ],
    "Summary": ["metric", "value"],
    "Methodology": ["section", "detail"],
}

SHEET_ORDER = [
    "Summary", "Weekly", "Quarterly", "Annual", "Unscheduled", "Watchlist",
    "Constituents", "FX", "Alerts", "Divisor_Log", "Audit", "Methodology",
]


class WorkbookManager:
    def __init__(self, path: str = config.WORKBOOK_PATH) -> None:
        self.path = path
        self.wb = self._open()

    def _open(self) -> Workbook:
        if os.path.exists(self.path):
            wb = load_workbook(self.path)
        else:
            wb = Workbook()
            # drop the default sheet
            default = wb.active
            wb.remove(default)
        self._ensure_sheets(wb)
        return wb

    def _ensure_sheets(self, wb: Workbook) -> None:
        for name in SHEET_ORDER:
            if name not in wb.sheetnames:
                ws = wb.create_sheet(title=name)
                self._write_header(ws, SCHEMAS[name])
            else:
                ws = wb[name]
                if ws.max_row == 0 or all(c.value is None for c in ws[1]):
                    self._write_header(ws, SCHEMAS[name])
        # order sheets
        wb._sheets.sort(key=lambda s: SHEET_ORDER.index(s.title)
                        if s.title in SHEET_ORDER else 999)

    @staticmethod
    def _write_header(ws, headers: list[str]) -> None:
        for col, head in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=head)
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"

    # --------------------------------------------------------------------- #
    def headers(self, sheet: str) -> list[str]:
        return SCHEMAS[sheet]

    def append_row(self, sheet: str, row: dict[str, Any]) -> None:
        ws = self.wb[sheet]
        headers = SCHEMAS[sheet]
        values = [row.get(h) for h in headers]
        ws.append(values)

    def read_rows(self, sheet: str) -> list[dict[str, Any]]:
        ws = self.wb[sheet]
        headers = SCHEMAS[sheet]
        out: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r is None or all(v is None for v in r):
                continue
            out.append({h: r[i] if i < len(r) else None
                        for i, h in enumerate(headers)})
        return out

    def has_week(self, iso_week: str) -> bool:
        """Idempotency guard: is there already a Weekly row for this ISO week?"""
        return any(str(row.get("iso_week")) == str(iso_week)
                   for row in self.read_rows("Weekly"))

    def delete_week(self, iso_week: str) -> int:
        """Remove Weekly rows for an ISO week (used by run_weekly --force)."""
        kept = [r for r in self.read_rows("Weekly")
                if str(r.get("iso_week")) != str(iso_week)]
        removed = len(self.read_rows("Weekly")) - len(kept)
        self.replace_sheet_rows("Weekly", kept)
        return removed

    def replace_sheet_rows(self, sheet: str, rows: list[dict[str, Any]]) -> None:
        """Clear a sheet's data rows (keep header) and rewrite them."""
        ws = self.wb[sheet]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for row in rows:
            self.append_row(sheet, row)

    def autosize(self) -> None:
        for name in self.wb.sheetnames:
            ws = self.wb[name]
            for col_cells in ws.columns:
                length = 0
                col = col_cells[0].column
                for cell in col_cells:
                    if cell.value is not None:
                        length = max(length, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col)].width = min(
                    max(length + 2, 10), 40)

    # --------------------------------------------------------------------- #
    # Summary / Methodology (human-facing dashboard sheets)
    # --------------------------------------------------------------------- #
    def write_summary(self, kpis: list[tuple[str, Any]],
                      top_weights: list[tuple[str, str, float]],
                      sectors: list[tuple[str, float]]) -> None:
        """Rewrite the Summary sheet: KPI block, top-10 weights, sector mix."""
        ws = self.wb["Summary"]
        ws.delete_rows(1, ws.max_row)
        bold = Font(bold=True)

        ws.cell(1, 1, "VGR Fashion 50 — Summary").font = Font(bold=True, size=14)
        r = 3
        ws.cell(r, 1, "Metric").font = bold
        ws.cell(r, 2, "Value").font = bold
        r += 1
        for k, v in kpis:
            ws.cell(r, 1, k)
            ws.cell(r, 2, v)
            r += 1

        r += 1
        ws.cell(r, 1, "Top 10 weights (post-cap)").font = bold
        r += 1
        ws.cell(r, 1, "ticker").font = bold
        ws.cell(r, 2, "name").font = bold
        ws.cell(r, 3, "weight_%").font = bold
        r += 1
        for tk, nm, w in top_weights:
            ws.cell(r, 1, tk)
            ws.cell(r, 2, nm)
            ws.cell(r, 3, round(w * 100, 4))
            r += 1

        r += 1
        ws.cell(r, 1, "Segment mix").font = bold
        r += 1
        ws.cell(r, 1, "segment").font = bold
        ws.cell(r, 2, "weight_%").font = bold
        r += 1
        for seg, w in sectors:
            ws.cell(r, 1, seg)
            ws.cell(r, 2, round(w * 100, 4))
            r += 1

    def add_index_chart(self) -> None:
        """Embed a native line chart of the Weekly index level on Summary."""
        weekly = self.wb["Weekly"]
        n = weekly.max_row
        if n < 3:
            return
        headers = SCHEMAS["Weekly"]
        lvl_col = headers.index("index_level") + 1
        date_col = headers.index("run_date") + 1

        chart = LineChart()
        chart.title = "VGR Fashion 50 — index level"
        chart.style = 2
        chart.y_axis.title = "Index (base 1000)"
        chart.x_axis.title = "Week"
        chart.height = 9
        chart.width = 24
        data = Reference(weekly, min_col=lvl_col, min_row=1, max_row=n)
        cats = Reference(weekly, min_col=date_col, min_row=2, max_row=n)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        # Anchor to the right of the KPI block.
        self.wb["Summary"].add_chart(chart, "E3")

    def write_prices(self, weeks, tickers, prices) -> None:
        """Write a wide weekly price panel: row = week, column = ticker.

        ``weeks`` list[date], ``tickers`` list[str], ``prices`` dict
        ticker -> {date: local_price}.  Backbone for the dashboard's per-company
        evolution selector and the YTD / MoM columns.
        """
        if "Prices" in self.wb.sheetnames:
            self.wb.remove(self.wb["Prices"])
        ws = self.wb.create_sheet("Prices")
        header = ["date"] + list(tickers)
        ws.append(header)
        for c in ws[1]:
            c.font = Font(bold=True)
        for w in weeks:
            ws.append([w.isoformat()]
                      + [prices.get(t, {}).get(w) for t in tickers])
        ws.freeze_panes = "B2"

    def write_subindices(self, dates, segments, levels) -> None:
        """Wide segment sub-index panel: row = date, column = segment level.

        ``dates`` list[str], ``segments`` list[str], ``levels`` dict
        segment -> {date_iso: level} (each rebased to 1000 at the base date).
        """
        if "SubIndices" in self.wb.sheetnames:
            self.wb.remove(self.wb["SubIndices"])
        ws = self.wb.create_sheet("SubIndices")
        ws.append(["date"] + list(segments))
        for c in ws[1]:
            c.font = Font(bold=True)
        for d in dates:
            ws.append([d] + [levels.get(s, {}).get(d) for s in segments])
        ws.freeze_panes = "B2"

    def write_methodology(self, sections: list[tuple[str, str]]) -> None:
        self.replace_sheet_rows(
            "Methodology",
            [{"section": s, "detail": d} for s, d in sections])

    def save(self) -> None:
        os.makedirs(os.path.dirname(self.path) or ".", exist_ok=True)
        self.autosize()
        self.wb.save(self.path)
